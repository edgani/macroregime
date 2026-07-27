from __future__ import annotations

import hashlib
import json
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import norm

ROOT = Path(__file__).resolve().parents[2]
PROTOCOL_PATH = ROOT / 'research_v78/protocols/V78_CROSS_MARKET_TSMOM_PROTOCOL_FROZEN.json'
RESULTS_PATH = ROOT / 'research_v78/results/V78_CROSS_MARKET_TSMOM_RESULTS.json'
LEDGER_PATH = ROOT / 'research_v78/ledgers/V78_CROSS_MARKET_TSMOM_LEDGER.csv'
SUMMARY_PATH = ROOT / 'research_v78/results/V78_CROSS_MARKET_TSMOM_SUMMARY.csv'


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_protocol() -> dict:
    p = json.loads(PROTOCOL_PATH.read_text())
    if p.get('status') != 'FROZEN_BEFORE_RESULT_COMPUTATION':
        raise RuntimeError('protocol is not frozen')
    source = ROOT / p['input']['path']
    if sha256(source) != p['input']['sha256']:
        raise RuntimeError('input hash does not match frozen protocol')
    return p


def load_factor_returns(path: Path, columns: list[str]) -> pd.DataFrame:
    ws = load_workbook(path, read_only=True, data_only=True)['TSMOM Factors']
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = list(row)
        if 'TSMOM' in vals:
            header_row = i
            headers = vals
            break
    if header_row is None:
        raise RuntimeError('TSMOM header not found')
    positions = {str(v): j for j, v in enumerate(headers) if v not in (None, '')}
    missing = [c for c in columns if c not in positions]
    if missing:
        raise RuntimeError(f'missing columns: {missing}')
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        dt = row[0]
        if not hasattr(dt, 'year'):
            continue
        rec = {'date': pd.Timestamp(dt)}
        valid = True
        for c in columns:
            v = row[positions[c]]
            if v in (None, ''):
                valid = False
                break
            rec[c] = float(v)
        if valid:
            rows.append(rec)
    df = pd.DataFrame(rows).set_index('date').sort_index()
    if df.empty or df.index.has_duplicates:
        raise RuntimeError('invalid factor panel')
    return df


def newey_west_se_mean(x: np.ndarray, lag: int = 6) -> float:
    x = np.asarray(x, dtype=float)
    x = x[np.isfinite(x)]
    n = len(x)
    if n < max(24, lag + 3):
        return float('nan')
    e = x - x.mean()
    gamma0 = float(np.dot(e, e) / n)
    lrv = gamma0
    for k in range(1, min(lag, n - 2) + 1):
        gamma = float(np.dot(e[k:], e[:-k]) / n)
        weight = 1.0 - k / (lag + 1.0)
        lrv += 2.0 * weight * gamma
    return math.sqrt(max(lrv, 0.0) / n)


def max_drawdown(r: np.ndarray) -> float:
    wealth = np.cumprod(1.0 + np.asarray(r, dtype=float))
    peak = np.maximum.accumulate(wealth)
    dd = wealth / peak - 1.0
    return float(np.nanmin(dd))


def moving_block_boot_prob(x: np.ndarray, hurdle: float, reps: int, block: int, seed: int) -> float:
    x = np.asarray(x, dtype=float)
    n = len(x)
    rng = np.random.default_rng(seed)
    starts = np.arange(n)
    means = np.empty(reps, dtype=float)
    nblocks = int(math.ceil(n / block))
    offsets = np.arange(block)
    for i in range(reps):
        s = rng.choice(starts, size=nblocks, replace=True)
        idx = ((s[:, None] + offsets[None, :]) % n).reshape(-1)[:n]
        means[i] = x[idx].mean()
    return float(np.mean(means > hurdle))


def period_slice(df: pd.DataFrame, spec: str) -> pd.DataFrame:
    start, end = spec.split('/')
    return df.loc[pd.Timestamp(start):pd.Timestamp(end)]


def evaluate_series(s: pd.Series, hurdle: float, critical_z: float, bootstrap: bool, seed: int) -> dict:
    x = s.dropna().to_numpy(float)
    n = len(x)
    mean = float(np.mean(x))
    se = newey_west_se_mean(x, lag=6)
    t = (mean - hurdle) / se if se > 0 else float('nan')
    lb = mean - critical_z * se - hurdle if se > 0 else float('nan')
    rolling = s.rolling(36).mean().dropna()
    years = s.groupby(s.index.year).mean()
    loo = []
    for y in years.index:
        sub = s[s.index.year != y]
        if len(sub):
            loo.append(float(sub.mean() - hurdle))
    reverse_mean = float((-s).mean())
    reverse_lb = reverse_mean - critical_z * se - hurdle if se > 0 else float('nan')
    out = {
        'n_months': n,
        'mean_monthly': mean,
        'annual_arithmetic_return': mean * 12.0,
        'annual_volatility': float(np.std(x, ddof=1) * math.sqrt(12.0)),
        'annual_sharpe': float(mean / np.std(x, ddof=1) * math.sqrt(12.0)) if np.std(x, ddof=1) > 0 else None,
        'max_drawdown': max_drawdown(x),
        'hurdle_monthly': hurdle,
        'hac_se_mean': se,
        'hac_t_above_hurdle': t,
        'bonferroni_adjusted_lower_bound_above_hurdle': lb,
        'rolling36_mean_above_hurdle_share': float((rolling > hurdle).mean()) if len(rolling) else None,
        'positive_calendar_year_share_above_hurdle': float((years > hurdle).mean()) if len(years) else None,
        'leave_one_year_out_positive_share_above_hurdle': float(np.mean(np.asarray(loo) > 0)) if loo else None,
        'reverse_sign_adjusted_lower_bound_above_hurdle': reverse_lb,
        'reverse_sign_pass': bool(reverse_lb > 0),
    }
    if bootstrap:
        out['moving_block_boot_probability_mean_above_hurdle'] = moving_block_boot_prob(x, hurdle, 20000, 12, seed)
    return out


def main() -> None:
    protocol = load_protocol()
    source = ROOT / protocol['input']['path']
    cols = protocol['candidate_family']
    df = load_factor_returns(source, cols)
    alpha = 0.05
    critical_z = float(norm.ppf(1.0 - alpha / len(cols)))
    hurdles = [float(x) for x in protocol['hurdles_monthly']]
    periods = protocol['periods']
    results = {
        'schema': 'warroom.v78.cross_market_tsmom_results.v1',
        'protocol_path': str(PROTOCOL_PATH.relative_to(ROOT)),
        'protocol_sha256': sha256(PROTOCOL_PATH),
        'input_sha256': sha256(source),
        'candidate_count': len(cols),
        'bonferroni_one_sided_critical_z': critical_z,
        'periods': {},
        'claim_boundary': protocol['claim_boundary'],
        'capital_permission': 'BLOCKED',
        'live_decision_weight': 0.0,
    }
    ledger = []
    summary = []
    for pidx, (pname, pspec) in enumerate(periods.items()):
        sub = period_slice(df, pspec)
        results['periods'][pname] = {'range': pspec, 'n_months': len(sub), 'series': {}}
        for cidx, c in enumerate(cols):
            cres = {}
            for h in hurdles:
                ev = evaluate_series(sub[c], h, critical_z, bootstrap=(h == 0.001), seed=7801 + pidx * 100 + cidx)
                cres[f'hurdle_{int(round(h*10000))}bp'] = ev
                ledger.append({
                    'period': pname, 'series': c, 'hurdle_monthly': h,
                    **{k: v for k, v in ev.items() if not isinstance(v, (dict, list))}
                })
            results['periods'][pname]['series'][c] = cres

    gate = protocol['promotion_gate_archive_supported']
    promoted = []
    for c in cols:
        v = results['periods']['post_publication_validation']['series'][c]['hurdle_10bp']
        l = results['periods']['post_publication_lockbox']['series'][c]['hurdle_10bp']
        passed = bool(
            v['bonferroni_adjusted_lower_bound_above_hurdle'] > 0 and
            l['bonferroni_adjusted_lower_bound_above_hurdle'] > 0 and
            v['moving_block_boot_probability_mean_above_hurdle'] >= gate['validation_block_boot_probability_mean_above_10bp_min'] and
            l['moving_block_boot_probability_mean_above_hurdle'] >= gate['lockbox_block_boot_probability_mean_above_10bp_min'] and
            v['rolling36_mean_above_hurdle_share'] >= gate['validation_rolling36_positive_share_min'] and
            l['rolling36_mean_above_hurdle_share'] >= gate['lockbox_rolling36_positive_share_min'] and
            v['positive_calendar_year_share_above_hurdle'] >= gate['validation_positive_calendar_year_share_min'] and
            l['positive_calendar_year_share_above_hurdle'] >= gate['lockbox_positive_calendar_year_share_min'] and
            not v['reverse_sign_pass'] and not l['reverse_sign_pass']
        )
        status = 'ARCHIVE_SUPPORTED_POST_PUBLICATION_10BP' if passed else 'NOT_PROMOTED'
        if passed:
            promoted.append(c)
        summary.append({
            'series': c,
            'status': status,
            'validation_mean_monthly': v['mean_monthly'],
            'validation_adjusted_lb_after_10bp': v['bonferroni_adjusted_lower_bound_above_hurdle'],
            'validation_boot_prob_after_10bp': v['moving_block_boot_probability_mean_above_hurdle'],
            'validation_rolling36_positive_share': v['rolling36_mean_above_hurdle_share'],
            'lockbox_mean_monthly': l['mean_monthly'],
            'lockbox_adjusted_lb_after_10bp': l['bonferroni_adjusted_lower_bound_above_hurdle'],
            'lockbox_boot_prob_after_10bp': l['moving_block_boot_probability_mean_above_hurdle'],
            'lockbox_rolling36_positive_share': l['rolling36_mean_above_hurdle_share'],
            'lockbox_25bp_adjusted_lb': results['periods']['post_publication_lockbox']['series'][c]['hurdle_25bp']['bonferroni_adjusted_lower_bound_above_hurdle'],
            'capital_permission': 'BLOCKED',
            'live_decision_weight': 0.0,
        })
    results['promotion'] = {
        'promoted_archive_supported_series': promoted,
        'promoted_count': len(promoted),
        'decision_active_count': 0,
        'reason_decision_inactive': 'Aggregate maintained factor returns do not provide exact instrument-level current positions or implementation-cost/capacity proof.'
    }
    RESULTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
    RESULTS_PATH.write_text(json.dumps(results, indent=2, sort_keys=True))
    pd.DataFrame(ledger).to_csv(LEDGER_PATH, index=False)
    pd.DataFrame(summary).to_csv(SUMMARY_PATH, index=False)
    print(json.dumps({
        'results': str(RESULTS_PATH),
        'ledger': str(LEDGER_PATH),
        'summary': str(SUMMARY_PATH),
        'promoted_archive_supported_series': promoted,
        'decision_active_count': 0,
    }, indent=2))


if __name__ == '__main__':
    main()
